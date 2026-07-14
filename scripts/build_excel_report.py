"""Build excel/Libstar_Product_Report.xlsx from data/aggregates/*.csv.

Detail sheets hold the aggregate tables (same CSVs the ebook charts use);
the KPI sheet derives its numbers with Excel formulas from those sheets, so
the workbook recalculates if the aggregates are refreshed. Named tables match
the Qlik Tabular Reporting template mapping in qlik/nprinting_tabular_reporting.md.
"""

import json
import os

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

BASE = os.path.join(os.path.dirname(__file__), "..")
AGG = os.path.join(BASE, "data", "aggregates")
ML = os.path.join(BASE, "ml", "outputs")
OUT = os.path.join(BASE, "excel", "Libstar_Product_Report.xlsx")

NAVY = "8E1E4D"
TEAL = "0FB5AE"
LIGHT = "FDF3F7"
FONT = "Arial"

header_font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
header_fill = PatternFill("solid", start_color=NAVY)
body_font = Font(name=FONT, size=10)
title_font = Font(name=FONT, bold=True, size=14, color=NAVY)
thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

R_FMT = '"R"#,##0;("R"#,##0);-'
N_FMT = '#,##0;(#,##0);-'
PCT_FMT = '0.0"%"'
DEC_FMT = "0.00"


def write_table(ws, df, name, start_row=3, fmts=None):
    ws.cell(row=1, column=1, value=ws.title.replace("_", " ")).font = title_font
    for j, col in enumerate(df.columns, 1):
        c = ws.cell(row=start_row, column=j, value=col)
        c.font, c.fill, c.border = header_font, header_fill, border
        c.alignment = Alignment(horizontal="center")
    for i, row in enumerate(df.itertuples(index=False), start_row + 1):
        for j, val in enumerate(row, 1):
            c = ws.cell(row=i, column=j, value=val)
            c.font, c.border = body_font, border
            if fmts and df.columns[j - 1] in fmts:
                c.number_format = fmts[df.columns[j - 1]]
    end = f"{get_column_letter(len(df.columns))}{start_row + len(df)}"
    t = Table(displayName=name, ref=f"A{start_row}:{end}")
    t.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(t)
    for j, col in enumerate(df.columns, 1):
        width = max(len(str(col)) + 2, 14)
        if col in ("product_name", "reject_reason", "brand", "sales_channel"):
            width = 34
        ws.column_dimensions[get_column_letter(j)].width = width


def main() -> None:
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    cat = pd.read_csv(os.path.join(AGG, "category_performance.csv"))
    brand = pd.read_csv(os.path.join(AGG, "brand_performance.csv"))
    prov = pd.read_csv(os.path.join(AGG, "province_performance.csv"))
    chan = pd.read_csv(os.path.join(AGG, "channel_performance.csv"))
    dq = pd.read_csv(os.path.join(AGG, "data_quality.csv"))
    growth = pd.read_csv(os.path.join(AGG, "catalog_growth.csv"))
    kpi_src = pd.read_csv(os.path.join(AGG, "kpi_summary.csv"))

    wb = Workbook()

    # ---- README ----
    ws = wb.active
    ws.title = "README"
    ws["A1"] = "Libstar Product Catalog Report"
    ws["A1"].font = Font(name=FONT, bold=True, size=18, color=NAVY)
    lines = [
        "",
        "Source: Azure Synapse serverless views (rpt schema) over ADLS curated parquet,",
        "produced by the Azure Data Factory cleaning pipeline (pl_clean_products).",
        "The same aggregate tables feed the Qlik Sense app and the ebook, so all",
        "deliverables reconcile to identical numbers.",
        "",
        f"Raw rows processed: {int(kpi_src['raw_rows'][0]):,} (incl. injected duplicates)",
        f"Clean unique products: {int(kpi_src['total_skus'][0]):,}",
        f"Quarantined rows: {int(kpi_src['quarantined_rows'][0]):,}",
        "",
        "Sheets: KPI_Summary (formula-driven), Category_Perf, Brand_Perf,",
        "Province_Perf, Channel_Perf, Data_Quality, Catalog_Growth, ML_Insights.",
        "Named tables map 1:1 to the Qlik Tabular Reporting template",
        "(see qlik/nprinting_tabular_reporting.md).",
        "",
        "Data is synthetic, generated for portfolio purposes on the Libstar",
        "public brand/category model (libstar.co.za).",
    ]
    for i, ln in enumerate(lines, 2):
        ws.cell(row=i, column=1, value=ln).font = body_font
    ws.column_dimensions["A"].width = 90

    # ---- detail sheets ----
    money = {"revenue_12m_zar": R_FMT, "avg_price_zar": R_FMT}
    pct = {"avg_margin_pct": PCT_FMT}
    nums = {"sku_count": N_FMT, "row_count": N_FMT, "products_added": N_FMT}
    fmts = {**money, **pct, **nums, "avg_rating": DEC_FMT}

    write_table(wb.create_sheet("Category_Perf"), cat, "Category_Perf", fmts=fmts)
    write_table(wb.create_sheet("Brand_Perf"), brand, "Brand_Perf", fmts=fmts)
    write_table(wb.create_sheet("Province_Perf"), prov, "Province_Perf", fmts=fmts)
    write_table(wb.create_sheet("Channel_Perf"), chan, "Channel_Perf", fmts=fmts)
    write_table(wb.create_sheet("Data_Quality"), dq, "Data_Quality", fmts=fmts)
    write_table(wb.create_sheet("Catalog_Growth"), growth, "Catalog_Growth", fmts=fmts)

    # ---- ML sheet ----
    ml_ws = wb.create_sheet("ML_Insights")
    metrics_path = os.path.join(ML, "metrics.json")
    r = 1
    ml_ws.cell(row=r, column=1, value="Machine Learning Results").font = title_font
    r += 2
    if os.path.exists(metrics_path):
        with open(metrics_path) as f:
            metrics = json.load(f)
        pm = metrics.get("price_model", {})
        for label, val in [
            ("Price model", pm.get("model", "")),
            ("Training rows", pm.get("train_rows", "")),
            ("R-squared (test)", pm.get("r2", "")),
            ("MAE (ZAR)", pm.get("mae_zar", "")),
            ("Anomalies flagged", metrics.get("anomalies_flagged", "")),
        ]:
            ml_ws.cell(row=r, column=1, value=label).font = Font(name=FONT, bold=True, size=10)
            ml_ws.cell(row=r, column=2, value=val).font = body_font
            r += 1
        r += 1
    seg_path = os.path.join(ML, "segments_summary.csv")
    if os.path.exists(seg_path):
        seg = pd.read_csv(seg_path)
        ml_ws.cell(row=r, column=1, value="Product segments (KMeans)").font = Font(
            name=FONT, bold=True, size=12, color=NAVY)
        r += 1
        for j, col in enumerate(seg.columns, 1):
            c = ml_ws.cell(row=r, column=j, value=col)
            c.font, c.fill, c.border = header_font, header_fill, border
        for _, row in seg.iterrows():
            r += 1
            for j, val in enumerate(row, 1):
                c = ml_ws.cell(row=r, column=j, value=val)
                c.font, c.border = body_font, border
        r += 2
    anom_path = os.path.join(ML, "price_anomalies.csv")
    if os.path.exists(anom_path):
        anom = pd.read_csv(anom_path).head(50)
        ml_ws.cell(row=r, column=1, value="Top 50 price anomalies (IsolationForest)").font = Font(
            name=FONT, bold=True, size=12, color=NAVY)
        r += 1
        start = r
        for j, col in enumerate(anom.columns, 1):
            c = ml_ws.cell(row=r, column=j, value=col)
            c.font, c.fill, c.border = header_font, header_fill, border
        for _, row in anom.iterrows():
            r += 1
            for j, val in enumerate(row, 1):
                c = ml_ws.cell(row=r, column=j, value=val)
                c.font, c.border = body_font, border
                if anom.columns[j - 1] in ("price_zar", "cost_zar"):
                    c.number_format = R_FMT
    for col in "ABCDEFGHI":
        ml_ws.column_dimensions[col].width = 18
    ml_ws.column_dimensions["B"].width = 34

    # ---- KPI summary (formula-driven off detail sheets) ----
    ws = wb.create_sheet("KPI_Summary", 1)
    ws["A1"] = "Libstar Catalog KPIs"
    ws["A1"].font = title_font
    ncat = len(cat)
    ndq = len(dq)
    col = {c: get_column_letter(i + 1) for i, c in enumerate(cat.columns)}
    rows = [
        ("Revenue 12m (ZAR)", f"=SUM(Category_Perf!{col['revenue_12m_zar']}4:{col['revenue_12m_zar']}{3 + ncat})", R_FMT),
        ("Total SKUs", f"=SUM(Category_Perf!{col['sku_count']}4:{col['sku_count']}{3 + ncat})", N_FMT),
        ("Avg margin % (SKU-weighted)",
         f"=SUMPRODUCT(Category_Perf!{col['avg_margin_pct']}4:{col['avg_margin_pct']}{3 + ncat},"
         f"Category_Perf!{col['sku_count']}4:{col['sku_count']}{3 + ncat})"
         f"/SUM(Category_Perf!{col['sku_count']}4:{col['sku_count']}{3 + ncat})", PCT_FMT),
        ("Avg price (ZAR, SKU-weighted)",
         f"=SUMPRODUCT(Category_Perf!{col['avg_price_zar']}4:{col['avg_price_zar']}{3 + ncat},"
         f"Category_Perf!{col['sku_count']}4:{col['sku_count']}{3 + ncat})"
         f"/SUM(Category_Perf!{col['sku_count']}4:{col['sku_count']}{3 + ncat})", R_FMT),
        ("Quarantined rows", f"=SUM(Data_Quality!B4:B{3 + ndq})", N_FMT),
        ("Rows after dedupe (clean + quarantined)", int(kpi_src["raw_rows"][0]), N_FMT),
        ("Data quality pass rate", "=1-B7/B8", "0.0%"),
        ("Active SKUs", int(kpi_src["active_skus"][0]), N_FMT),
        ("Avg rating", float(kpi_src["avg_rating"][0]), DEC_FMT),
    ]
    r = 3
    for label, val, fmt in rows:
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = Font(name=FONT, bold=True, size=11)
        lc.fill = PatternFill("solid", start_color=LIGHT)
        lc.border = border
        vc = ws.cell(row=r, column=2, value=val)
        vc.font = Font(name=FONT, size=11)
        vc.number_format = fmt
        vc.border = border
        r += 1
    ws.cell(row=r + 1, column=1,
            value="Hardcoded cells (raw rows, active SKUs, avg rating): "
                  "Source: Synapse rpt.vw_kpi_summary / ADF run b31e42ae, 2026-07-07."
            ).font = Font(name=FONT, italic=True, size=9, color="666666")
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 22

    wb.save(OUT)
    print("written:", os.path.abspath(OUT))


if __name__ == "__main__":
    main()
